import openpyxl,sys,os
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

def app_path():
    if hasattr(sys,'frozen'):
        return os.path.dirname(sys.executable)
    return os.path.dirname(__file__)

def insert(result_list,title):
    wb=openpyxl.Workbook()
    ws=wb.active

    ws['A1']='分类'
    ws['B1']='地址'
    ws['C1']='标题'
    ws['D1']='数量'
    for result in result_list:
        result1 = [ILLEGAL_CHARACTERS_RE.sub(r'', str(i)) for i in result]
        result1[-1]=int(result1[-1])
        ws.append(result1)
    path=app_path()+'\\'+title+'.xlsx'
    wb.save(path)
